"""
Generate sample Excel files for POC testing.
Run this script once to create 2 test Excel files with some overlapping phone numbers
to demonstrate the duplicate detection / merge logic.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

HEADERS = ["Lead Source (d)", "Employee NAme", "Customer Name", "Address", "Phone", "Status", "Remarks"]

SAMPLE1 = [
    ["Whatsapp", "Ravi Kumar", "Arjun Sharma",   "123 MG Road, Bengaluru",    "+91 9876543210", "First Call",               "Interested in home loan"],
    ["Website",  "Priya S",   "Sunita Devi",     "45 Anna Nagar, Chennai",    "9123456780",     "Followup Call",            "Called back, needs more info"],
    ["Justdial", "Ravi Kumar", "Mohammed Ali",   "7 Linking Road, Mumbai",    "0-9988776655",   "Rate Call",                "Asked for interest rates"],
    ["Walkin",   "Meena T",   "Kavya Reddy",     "Plot 9, Hyderabad",         "+919845001234",  "Assigned to Branch",       "Assigned to Koramangala"],
    ["TeleCall", "Priya S",   "Deepak Mehta",    "88 Sector 15, Noida",       "9001234567",     "Customer Not Interested",  "Not looking right now"],
    ["Whatsapp", "Ravi Kumar", "Latha Nair",     "12 Connaught Place, Delhi", "9912300077",     "Appointment Call",         "Meeting scheduled 10am Mon"],
]

SAMPLE2 = [
    # These two phones are duplicates (already in sample1 - different format)
    ["Website",  "New Agent",  "Arjun Sharma",   "Duplicate entry test",      "9876543210",     "Followup Call",  "Should be skipped"],
    ["Justdial", "New Agent",  "Mohammed Ali",   "Duplicate test 2",          "91 9988776655",  "First Call",     "Should also be skipped"],
    # These are NEW leads (unique phones)
    ["Walkin",   "Suresh B",   "Ananya Iyer",    "55 Brigade Road, Bengaluru","9700112233",     "First Call",               "Walk-in customer"],
    ["TeleCall", "Meena T",    "Rohit Sinha",    "22 Park Street, Kolkata",   "+91-9433000111", "Rate Call",                "Requested loan calculator"],
    ["Website",  "Suresh B",   "Fatima Begum",   "Apartment 4, Pune",         "020-9527001122", "Customer Interested Need Time", "Will call back in 2 weeks"],
    ["Whatsapp", "Priya S",    "Vikas Joshi",    "Navi Mumbai",               "9820099000",     "Followup Call",            "Second follow-up done"],
]

SUMMARY_ROW = ["Total Calls", "First Call Completed", "Followup Call", "Rate Call", "Assigned to branch", None, None]
SUMMARY_VALS = [10, 6, 2, 3, 1, None, None]


def style_header(cell):
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="4B4FBF")
    cell.alignment = Alignment(horizontal="center")


def create_sample_excel(filename, rows, title="Sample Leads"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Headers
    for col_idx, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        style_header(cell)
        ws.column_dimensions[cell.column_letter].width = 22

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Blank rows
    blank_end = len(rows) + 4

    # Summary row (like the reference Excel)
    sum_row = blank_end + 1
    for col_idx, val in enumerate(SUMMARY_ROW, 1):
        ws.cell(row=sum_row, column=col_idx, value=val).font = Font(bold=True)
    sum_row2 = sum_row + 1
    for col_idx, val in enumerate(SUMMARY_VALS, 1):
        ws.cell(row=sum_row2, column=col_idx, value=val)

    wb.save(filename)
    print(f"Created: {filename}")


if __name__ == "__main__":
    out_dir = os.path.join(os.path.dirname(__file__), "..", "sample_data")
    os.makedirs(out_dir, exist_ok=True)

    create_sample_excel(os.path.join(out_dir, "leads_batch_1.xlsx"), SAMPLE1)
    create_sample_excel(os.path.join(out_dir, "leads_batch_2.xlsx"), SAMPLE2)

    print("\nDone! 2 sample files created in sample_data/")
    print("leads_batch_1.xlsx - 6 fresh leads")
    print("leads_batch_2.xlsx - 2 DUPLICATES + 4 new leads")
    print("\nUpload batch_1 first, then batch_2 to see merge logic in action!")
