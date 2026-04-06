from fastapi import FastAPI, UploadFile, File, Depends, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, text
from datetime import date as date_type
import shutil
import os
import io
from datetime import datetime
from typing import List, Optional

from database import create_tables, get_db, Lead, UploadLog
from excel_processor import read_excel_file
from phone_utils import normalize_phone

app = FastAPI(title="Lead Management POC", version="1.0.0")

# CORS - allow all origins for POC
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Create tables on startup
create_tables()

# Serve frontend static files
FRONTEND_DIR = os.path.join(os.path.dirname(__file__), "..", "frontend")
if os.path.exists(FRONTEND_DIR):
    app.mount("/static", StaticFiles(directory=FRONTEND_DIR), name="static")


@app.get("/")
def serve_frontend():
    index_path = os.path.join(FRONTEND_DIR, "index.html")
    if os.path.exists(index_path):
        return FileResponse(index_path)
    return {"message": "Lead Management API is running. Frontend not found."}


@app.post("/api/upload")
async def upload_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """
    Upload an Excel file, process it, and merge into the database.
    Duplicates are detected by normalized phone numbers.
    """
    # Validate file type
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx and .xls files are supported.")

    # Read file entirely into memory (BytesIO) — avoids ALL Windows file-locking issues
    try:
        contents = await file.read()
        file_stream = io.BytesIO(contents)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to read uploaded file: {str(e)}")

    # Parse the Excel from memory
    try:
        leads_data = read_excel_file(file_stream)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Error reading Excel file: {str(e)}")


    inserted = 0
    duplicates = 0
    errors = 0

    for lead_dict in leads_data:
        try:
            phone_raw = lead_dict.get("phone") or ""
            
            # Skip if phone is empty (Manager's requirement)
            if not phone_raw:
                errors += 1
                continue
                
            phone_normalized = normalize_phone(phone_raw)

            # Duplicate check: if phone is present, check if it already exists
            is_duplicate = False
            if phone_normalized:
                existing = db.query(Lead).filter(Lead.phone == phone_normalized).first()
                if existing:
                    is_duplicate = True
                    duplicates += 1

            if not is_duplicate:
                now = datetime.utcnow()
                new_lead = Lead(
                    lead_source=lead_dict.get("lead_source"),
                    employee_name=lead_dict.get("employee_name"),
                    customer_name=lead_dict.get("customer_name"),
                    address=lead_dict.get("address"),
                    phone=phone_normalized if phone_normalized else phone_raw,
                    phone_raw=phone_raw,
                    status=lead_dict.get("status"),
                    remarks=lead_dict.get("remarks"),
                    source_file=file.filename,
                    lead_date=now,          # date this lead was uploaded/recorded
                    uploaded_at=now,
                )
                db.add(new_lead)
                inserted += 1
        except Exception as e:
            errors += 1
            continue

    db.commit()

    # Save upload log
    log = UploadLog(
        filename=file.filename,
        total_rows=len(leads_data),
        inserted=inserted,
        duplicates=duplicates,
        errors=errors,
    )
    db.add(log)
    db.commit()

    return {
        "success": True,
        "filename": file.filename,
        "total_rows": len(leads_data),
        "inserted": inserted,
        "duplicates_skipped": duplicates,
        "errors": errors,
        "message": f"✅ Processed {len(leads_data)} rows: {inserted} inserted, {duplicates} duplicates skipped, {errors} errors.",
    }


@app.get("/api/leads")
def get_leads(
    skip: int = 0,
    limit: int = 200,
    status: Optional[str] = None,
    lead_source: Optional[str] = None,
    search: Optional[str] = None,
    date_from: Optional[str] = None,   # format: YYYY-MM-DD
    date_to: Optional[str] = None,     # format: YYYY-MM-DD
    sort_order: Optional[str] = "desc",  # "asc" or "desc"
    db: Session = Depends(get_db),
):
    """Get paginated list of leads with filters and date-wise sorting."""
    query = db.query(Lead)

    if status:
        query = query.filter(Lead.status.ilike(f"%{status}%"))
    if lead_source:
        query = query.filter(Lead.lead_source.ilike(f"%{lead_source}%"))
    if search:
        query = query.filter(
            (Lead.customer_name.ilike(f"%{search}%")) |
            (Lead.phone.ilike(f"%{search}%")) |
            (Lead.employee_name.ilike(f"%{search}%"))
        )

    # Date range filter (using lead_date / uploaded_at)
    if date_from:
        try:
            dt_from = datetime.strptime(date_from, "%Y-%m-%d")
            query = query.filter(Lead.uploaded_at >= dt_from)
        except ValueError:
            pass
    if date_to:
        try:
            dt_to = datetime.strptime(date_to, "%Y-%m-%d")
            # Include the full day
            dt_to = dt_to.replace(hour=23, minute=59, second=59)
            query = query.filter(Lead.uploaded_at <= dt_to)
        except ValueError:
            pass

    total = query.count()

    # Sort by uploaded_at date
    if sort_order == "asc":
        leads = query.order_by(Lead.uploaded_at.asc()).offset(skip).limit(limit).all()
    else:
        leads = query.order_by(Lead.uploaded_at.desc()).offset(skip).limit(limit).all()

    return {
        "total": total,
        "leads": [
            {
                "id": l.id,
                "lead_source": l.lead_source,
                "employee_name": l.employee_name,
                "customer_name": l.customer_name,
                "address": l.address,
                "phone": l.phone,
                "phone_raw": l.phone_raw,
                "status": l.status,
                "remarks": l.remarks,
                "source_file": l.source_file,
                "lead_date": l.lead_date.strftime("%Y-%m-%d") if l.lead_date else None,
                "uploaded_at": l.uploaded_at.isoformat() if l.uploaded_at else None,
            }
            for l in leads
        ],
    }


@app.get("/api/dashboard")
def get_dashboard(
    date_from: str = Query(None),
    date_to: str = Query(None),
    db: Session = Depends(get_db)
):
    """Get dashboard statistics filtered by date range."""
    # Shared filter logic for dates
    def apply_date_filters(q, model):
        if date_from:
            try:
                dt_from = datetime.strptime(date_from, "%Y-%m-%d")
                q = q.filter(model.uploaded_at >= dt_from)
            except ValueError:
                pass
        if date_to:
            try:
                dt_to = datetime.strptime(date_to, "%Y-%m-%d")
                dt_to = dt_to.replace(hour=23, minute=59, second=59)
                q = q.filter(model.uploaded_at <= dt_to)
            except ValueError:
                pass
        return q

    # Total leads query with date filters
    total_q = db.query(Lead)
    total_q = apply_date_filters(total_q, Lead)
    total_leads = total_q.count()

    # Status breakdown with date filters
    status_q = db.query(Lead.status, func.count(Lead.id)).filter(Lead.status != None)
    status_q = apply_date_filters(status_q, Lead)
    status_counts = status_q.group_by(Lead.status).all()

    # Lead source breakdown with date filters
    source_q = db.query(Lead.lead_source, func.count(Lead.id)).filter(Lead.lead_source != None)
    source_q = apply_date_filters(source_q, Lead)
    source_counts = source_q.group_by(Lead.lead_source).all()

    # Employee breakdown with date filters
    employee_q = db.query(Lead.employee_name, func.count(Lead.id)).filter(Lead.employee_name != None)
    employee_q = apply_date_filters(employee_q, Lead)
    employee_counts = employee_q.group_by(Lead.employee_name).all()

    # Upload history doesn't usually filter by specific date range (shows last 10)
    upload_logs = (
        db.query(UploadLog)
        .order_by(UploadLog.uploaded_at.desc())
        .limit(10)
        .all()
    )

    return {
        "total_leads": total_leads,
        "status_breakdown": {s: c for s, c in status_counts},
        "source_breakdown": {s: c for s, c in source_counts},
        "employee_breakdown": {e: c for e, c in employee_counts},
        "upload_history": [
            {
                "filename": ul.filename,
                "total_rows": ul.total_rows,
                "inserted": ul.inserted,
                "duplicates": ul.duplicates,
                "errors": ul.errors,
                "uploaded_at": ul.uploaded_at.isoformat() if ul.uploaded_at else None,
            }
            for ul in upload_logs
        ],
    }


@app.delete("/api/leads/all")
def clear_all_leads(db: Session = Depends(get_db)):
    """Clear all leads (for testing POC)."""
    db.query(Lead).delete()
    db.query(UploadLog).delete()
    db.commit()
    return {"message": "All leads cleared."}


@app.get("/api/export")
def export_merged_excel(db: Session = Depends(get_db)):
    """
    Export ALL leads from the database as a single merged Excel file.
    This is the 'master merged file' — every upload contributes to it.
    Duplicates were already prevented at insert time.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    leads = db.query(Lead).order_by(Lead.uploaded_at.asc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Merged Leads"

    # -- Header styling --
    HEADER_FILL   = PatternFill("solid", fgColor="4B4FBF")
    HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
    ALT_FILL      = PatternFill("solid", fgColor="F0F0F8")
    CENTER        = Alignment(horizontal="center", vertical="center")
    THIN_BORDER   = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    headers = [
        "#", "Lead Source", "Employee Name", "Customer Name",
        "Address", "Phone", "Status", "Remarks", "Lead Date", "Source File"
    ]

    # Write headers
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # Freeze top row
    ws.freeze_panes = "A2"

    # Write data rows
    for row_idx, lead in enumerate(leads, 2):
        alt = row_idx % 2 == 0
        row_data = [
            row_idx - 1,
            lead.lead_source,
            lead.employee_name,
            lead.customer_name,
            lead.address,
            lead.phone,
            lead.status,
            lead.remarks,
            lead.lead_date.strftime("%d-%m-%Y") if lead.lead_date else None,
            lead.source_file,
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
            if alt:
                cell.fill = ALT_FILL

    # Auto-fit column widths
    col_widths = [5, 15, 18, 20, 30, 15, 28, 30, 14, 20]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 22

    # -- Summary section at bottom --
    summary_row = len(leads) + 3
    ws.cell(row=summary_row, column=1, value="SUMMARY").font = Font(bold=True, size=12)

    # Status counts
    from collections import Counter
    status_counts = Counter(l.status for l in leads if l.status)
    ws.cell(row=summary_row + 1, column=1, value="Status").font = Font(bold=True)
    ws.cell(row=summary_row + 1, column=2, value="Count").font = Font(bold=True)
    for i, (status, count) in enumerate(status_counts.items(), summary_row + 2):
        ws.cell(row=i, column=1, value=status)
        ws.cell(row=i, column=2, value=count)

    ws.cell(row=summary_row, column=4, value=f"Total Leads: {len(leads)}").font = Font(bold=True, size=12)

    # Stream the workbook from memory (no disk write)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from datetime import date
    filename = f"merged_leads_{date.today().strftime('%d_%m_%Y')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@app.get("/api/health")
def health_check():
    return {"status": "ok", "time": datetime.utcnow().isoformat()}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
