import re
import openpyxl
from typing import List, Dict, Any


# Status values that are valid (from dropdown column F)
VALID_STATUSES = {
    "first call", "followup call", "rate call", "assigned to branch",
    "appointment call", "customer not interested",
    "customer interested need time", "customer dropped"
}

# Lead sources
VALID_SOURCES = {
    "whatsapp", "website", "justdial", "walkin", "telecall"
}


def read_excel_file(file_source) -> List[Dict[str, Any]]:
    import io

    if isinstance(file_source, io.BytesIO):
        file_source.seek(0)

    try:
        wb = openpyxl.load_workbook(file_source, data_only=True)
        sheet = wb.active
    except Exception as e:
        raise ValueError(f"Cannot read Excel file: {e}. Ensure it is .xlsx format.")

    # Find the header row by looking for "Phone" or "Customer" keyword
    header_row_idx = None
    headers = []
    
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
        row_str = [str(v).strip().lower() for v in row if v is not None]
        if any("phone" in cell or "customer" in cell or "employee" in cell for cell in row_str):
            header_row_idx = row_idx
            headers = [str(v).strip().lower().replace(" ", "_").replace("-", "_") if v else "" for v in row]
            break

    if header_row_idx is None:
        header_row_idx = 1
        # Fallback empty headers
        headers = []

    # Map to standard field names
    col_map = {
        "lead_source": find_col(headers, ["lead source", "lead_source", "source"]),
        "employee_name": find_col(headers, ["employee name", "employee_name", "employee"]),
        "customer_name": find_col(headers, ["customer name", "customer_name", "customer"]),
        "address": find_col(headers, ["address", "addr"]),
        "phone": find_col(headers, ["phone", "mobile", "contact", "phone no", "phone number"]),
        "status": find_col(headers, ["status"]),
        "remarks": find_col(headers, ["remarks", "remark", "notes", "note"]),
    }

    leads = []
    for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
        lead = {}
        for field, col_idx in col_map.items():
            if col_idx is not None and col_idx < len(row):
                val = row[col_idx]
                lead[field] = clean_value(val)
            else:
                lead[field] = None

        # Skip summary / empty rows
        if is_lead_row_valid(lead):
            leads.append(lead)

    return leads


def is_lead_row_valid(lead: Dict) -> bool:
    """A valid lead row MUST have a phone number."""
    phone = str(lead.get("phone") or "").strip()
    lead_source = str(lead.get("lead_source") or "").strip().lower()

    # Skip pure summary or blank rows
    if "total" in lead_source or "calls" in lead_source:
        return False
    
    # NEW LOGIC: If no phone number, it's not a valid lead row for the DB
    if not phone or phone.lower() in ("nan", "none", "null"):
        return False
        
    return True


def find_col(headers: list, candidates: list):
    """Find a matching column index from a list of candidates."""
    for c in candidates:
        candidate_norm = c.replace(" ", "_")
        # Exact match
        if candidate_norm in headers:
            return headers.index(candidate_norm)
        # Partial match
        for idx, h in enumerate(headers):
            if candidate_norm in h or h in candidate_norm:
                return idx
    return None


def clean_value(val) -> str:
    if val is None:
        return None
    val = str(val).strip()
    if val.lower() in ("nan", "none", "null", ""):
        return None
    # Strip trailing .0 from numeric values
    if re.match(r"^\d+\.0$", val):
        val = val[:-2]
    return val
